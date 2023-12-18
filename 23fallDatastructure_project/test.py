# -*- coding: utf-8 -*-
from openpyxl import Workbook

import pandas as pd
import hashlib
import numpy as np
from queue import Queue,LifoQueue,PriorityQueue
import tkinter as tk
from tkinter import *
import tkinter
# 引入字体模块
from PIL import Image
from PIL import ImageTk
import tkinter.font as tkFont
import os

from openpyxl.reader.excel import load_workbook

a = np.zeros(10000, dtype='float32')  #初始化一个10000大小的数组，全赋值0,     储存数组
b = np.zeros(10000, dtype='float32')  #初始化一个10000大小的数组，全赋值0，    检测数组
q = Queue(maxsize=0)             #定义一个队列，不限制大小
q1= Queue(maxsize=0)
combined_data = []
global s1


def hash(name,score):
    s1=''    #储存字符串
    s=hashlib.md5(name.encode(encoding='UTF-8')).hexdigest()#md5将中文转字母+数字
    #print(s)   #测试输出用
    for ch in s:  # 第一个实例
        s1=s1+str(ord(ch)-40)   #移位加密，0的ascii码是48
    #print(s1)   #测试输出用
    pos=int(s1)%9973  #hash表的质数取余法 9973为一个接近10000的大质数
    print(name,pos)   #测试哈希值用
    if a[pos]==0:
        a[pos]=score
    else:
        while a[pos]!=0:
            pos=pos+1
        a[pos]=score
        #print(name,"出现哈希冲突")
        #print("一旦出现哈希冲突，我们便认为这个算法是不够安全的，在实际应用中建议将冲突学生姓名修改成 ”", m, "1“ 便可以保证算法的安全性")

def read(s,s1,m,y1,y2):
    # 行号和列号必须从0开始，获取指定单元格的内容
    df = pd.read_excel(s1)  # 打开文件
    while 1:
        try:
            name = df.iat[m, y1]  # 读取第m行第六列的值，这里不需要嵌套列表
            score = df.iat[m, y2]  # 读取第m行第零列的值，这里不需要嵌套列表
        except Exception as e:  #这里解决了一个文件指针的越界异常，一旦越界就返回菜单
            break

        if name == '':
            break
        hash(name, score)
        # print(name,score)
        m = m + 1

def conflict(s,s1):
    m = 0  # 行坐标
    flag=0
    # 行号和列号必须从0开始，获取指定单元格的内容

    df1 = pd.read_excel(s)  # 打开文件

    while 1:
        # 行号和列号必须从0开始，获取指定单元格的内容
        f3 = open("conflict.in", "w", encoding='UTF-8')  # 和上面的df1冲突，很离谱，无法解决
        try:
            namelist = df1.iat[m, 0]  #姓名顶在（0，0）开始
        except Exception as e:   #这里解决了一个文件指针的越界异常，一旦越界就返回菜单
            print(s,s1)
            break

        if namelist == '':
            break
        m = m + 1  #进到下一行
        s2= ''  # 储存字符串
        s3 = hashlib.md5(namelist.encode(encoding='UTF-8')).hexdigest()  # md5将中文转字母+数字
        # print(s)   #测试输出用
        for ch in s3:  # 第一个实例
            s2 = s2 + str(ord(ch) - 40)  # 移位加密，0的ascii码是48
        # print(s1)   #测试输出用
        pos = int(s2) % 9973  # hash表的质数取余法 9973为一个接近10000的大质数
        #print(namelist,pos)  # 测试哈希值用
        if b[pos] == 0:
            b[pos] = 1
        else:
            print(namelist, "出现哈希冲突")
            f3.write(namelist)
            f3.write("出现哈希冲突")
    f3.close()
    window2 = tk.Toplevel(win)
    window2.configure(bg='LightYellow')
    window2.title("冲突检测")
    window2.geometry("500x250")
    #Label(window2, text="请选择要执行的功能:", font=ft4).place(x=50, y=0)
    f2 = open("conflict.in", "r",encoding='UTF-8')
    code = f2.read()
    f2.close()
    text = tk.Text(window2)
    text.insert('end', code)
    text.place(x=0, y=60)

def output(s2,s3):
    m = 0  # 行坐标
    # 行号和列号必须从0开始，获取指定单元格的内容

    df1 = pd.read_excel(s2)  # 打开文件
    while 1:
        # 行号和列号必须从0开始，获取指定单元格的内容
        try:
            namelist = df1.iat[m, 0]
        except Exception as e:  #这里解决了一个文件指针的越界异常，一旦越界就返回菜单
            break
        if namelist == '':
            break
        m = m + 1  #进到下一行
        s1 = ''  # 储存字符串
        s = hashlib.md5(namelist.encode(encoding='UTF-8')).hexdigest()  # md5将中文转字母+数字
        # print(s)   #测试输出用
        for ch in s:  # 第一个实例
            s1 = s1 + str(ord(ch) - 40)  # 移位加密，0的ascii码是48
        # print(s1)   #测试输出用
        pos = int(s1) % 9973  # hash表的质数取余法 9973为一个接近10000的大质数
        #print(namelist,pos)  # 测试哈希值用
        if a[pos] != -1:
            print(namelist,a[pos])  #哈希表删除时不能将该点元素删除，否则查找不到后面的冲突元素，一般置一个取不到的特殊符号，这里置-1
            q.put(a[pos])      #将成绩入队
            a[pos]=-1
        else:
            while a[pos]==-1:
                pos=pos+1
            print(namelist, a[pos])  # 哈希表删除时不能将该点元素删除，否则查找不到后面的冲突元素，一般置一个取不到的特殊符号，这里置-1
            q.put(a[pos])
            a[pos] = -1
    output1(s)

def output1(s):

    # with open("sorted_scores.out", "w", encoding='UTF-8') as f_out:
    #     f_out.write('\n'.join(combined_data))
    score_sort()
    # Rest of your code for tkinter window...
    window3 = tk.Toplevel(win)
    window3.configure(bg='LightYellow')
    window3.title("成绩预览")
    window3.geometry("800x550")

    f_out = open("sorted_scores.out", "r", encoding='UTF-8')
    code1 = f_out.read()
    f_out.close()

    text1 = tk.Text(window3, width=100, height=30)
    text1.insert('end', code1)
    text1.place(x=10, y=60)

def get_xy(temp):
    temp=temp.upper()  #全转大写
    y=ord(temp)-65
    return int(y)      #要转int

def menu1(s,s1):

    print("1.使用说明")
    print("2.哈希冲突检测")
    print("3.成绩文件读取")
    print("4.预览学生名单输出")
    print("5.获取输出的成绩")
    print("0.退出程序")


    x = input("请选择执行功能：")

    if x == '1':
        print("首先执行2.哈希冲突检测，如果未发生冲突执行3.成绩文件读取，读取之后可以执行4.学生名单成绩输出")
        menu(s,s1)
    if x == '2':
        print("一旦出现哈希冲突，我们便认为这个算法是不够安全的，在实际应用中建议将冲突学生姓名修改成 ”张三1“ 便可以保证算法的安全性")
        conflict(s,s1)
        menu(s,s1)
    if x == '3':
        m=0    #行坐标
        y1=1   #名字的起始列坐标
        y2=4   #成绩的起始列坐标
        print("默认参数为名字起始坐标(2,B),成绩起始坐标（2,E)，注意起始行坐标要在同一行")
        c = input("如果要修改默认参数选择1，否则选择0：")
        if c=='1':     #注意直接输入的是char
            m = input("请输入起始名字的行坐标：")
            temp = input("请输入起始名字的列坐标：")
            y1= get_xy(temp)
            m = input("请输入起始成绩的行坐标：")
            temp = input("请输入起始成绩的列坐标：")
            y2 = get_xy(temp)
            m=int(m)    #注意要转int
            read(s,s1, m, y1, y2)
        else:
            read(s,s1, m, y1, y2)
        menu(s,s1)
    if x == '4':
        output(s,s1)
        menu(s,s1)
    if x == '5':
        output1(s)
        menu(s,s1)
    if x == '0':
        exit(0)
def name(s):
    m = 0  # 行坐标
    df3 = pd.read_excel(s)  # 打开文件

    while 1:
        # 行号和列号必须从0开始，获取指定单元格的内容
        #f3 = open("name.in", "w", encoding='UTF-8')  # 和上面的df1冲突，很离谱，无法解决
        try:
            namelist = df3.iat[m, 0]  # 姓名顶在（0，0）开始
        except Exception as e:  # 这里解决了一个文件指针的越界异常，一旦越界就返回菜单
            break

        if namelist == '':
            break

        q1.put(namelist)
        m = m + 1  # 进到下一行
    wr()

def wr():
    f1 = open("name.in", "w")
    while 1:
        if q1.empty():  #检验队列是否为空
            break
        #print(q1.get())  #不为空输出队首且将队首删除
        f1.write(str(q1.get()))
        f1.write('\n')
    f1.close()
def image(window_creat):
    #load = Image.open("1.jpg")  # open image from path
    #img = ImageTk.PhotoImage(load)  # read opened image

    #label1 = tkinter.Label(window_creat, image=img).place(x=500, y=0)  # create a label to insert this image
    #label1.grid()  # set the label in the main window
    img = PhotoImage(file="2.gif")
    label4 = Label(window_creat, image=img, text="111", compound="bottom", bg="lightgray").place(x=500, y=0)
    label4.grid()

# def edit():
#     window_edit = tk.Toplevel(win)
#     window_edit.configure(bg='LightBlue')
#     window_edit.title("学生信息修改界面")
#     window_edit.geometry("600x500")
#
#     Label(window_edit, text="请选择要执行的功能:", bg='LightYellow', fg='Tomato',font=ft4).place(x=50, y=0)
#
#     var_edit = StringVar()
#     Label_edit = tk.Label(win, text="", bg='LightGreen', fg='Maroon', justify='left').place(x=50, y=50)
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def copy_sheet(source_file, source_sheet_name, destination_file, destination_sheet_name):
    # 打开源文件和目标文件
    source_wb = load_workbook(source_file)
    destination_wb = Workbook()

    # 获取源表格和目标表格
    source_sheet = source_wb[source_sheet_name]
    destination_sheet = destination_wb.active

    # 复制源表格的数据到目标表格
    for row in source_sheet.iter_rows():
        for cell in row:
            destination_sheet[cell.coordinate].value = cell.value

    # 更改目标表格的名称
    destination_sheet.title = destination_sheet_name

    # 保存目标文件
    destination_wb.save(destination_file)
    destination_wb.close()

# copy_sheet('source.xlsx', 'Sheet1', 'destination.xlsx', 'NewSheet')

def edit(s,s1):
    print(s1)
    # copy_sheet(s1,'Sheet1','e.xlsx','Sheet1')
    # tt='e.xlsx'
    wb = load_workbook(s1)  # 修改成绩文件
    wb1= load_workbook(s)   # 修改名单文件
    print('1111')
    def submit_info(wb,wb1,s1,s):
        # 获取输入框中的信息
        name = entry_name.get()
        student_id = entry_id.get()
        score = entry_score.get()
        flag = int(entry_flag.get())
        print(flag)
        # 打开 Excel 文件
        # wb = load_workbook('e.xlsx')
        ws = wb.active
        ws1= wb1.active

        if flag==1:
        # 将信息添加到 Excel 的最后一行
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=2).value = name
            ws.cell(row=last_row, column=3).value = student_id
            ws.cell(row=last_row, column=4).value = score

            last_row1=ws1.max_row+1
            ws1.cell(row=last_row, column=1).value = name
        elif flag==2:
            print("here")
            for row in ws.iter_rows(min_row=1, min_col=2,max_col=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value == name:
                        ws.delete_rows(cell.row, 1)  # 删除找到的行
            for row in ws1.iter_rows(min_row=1, min_col=1,max_col=1, max_row=ws1.max_row):
                for cell in row:
                    if cell.value == name:
                        ws1.delete_rows(cell.row, 1)  # 删除找到的行
        # 保存并关闭 Excel 文件
        wb.save(s1)
        wb.close()
        wb1.save(s)
        wb1.close()

    window_edit = tk.Toplevel(win)
    window_edit.configure(bg='LightBlue')
    window_edit.title("学生信息修改界面")
    window_edit.geometry("600x500")

    Label(window_edit, text="请输入以下信息:", bg='LightYellow', fg='Tomato', font=ft4).place(x=50, y=0)

    # 创建输入框和标签
    Label(window_edit, text="添加or删除信息：输入1or2", bg='LightGreen', fg='Maroon').place(x=400, y=70)
    entry_flag = tk.Entry(window_edit)
    entry_flag.place(x=400, y=100)

    Label(window_edit, text="姓名:", bg='LightGreen', fg='Maroon').place(x=50, y=50)
    entry_name = tk.Entry(window_edit)
    entry_name.place(x=150, y=50)

    Label(window_edit, text="学号:", bg='LightGreen', fg='Maroon').place(x=50, y=100)
    entry_id = tk.Entry(window_edit)
    entry_id.place(x=150, y=100)

    Label(window_edit, text="成绩:", bg='LightGreen', fg='Maroon').place(x=50, y=150)
    entry_score = tk.Entry(window_edit)
    entry_score.place(x=150, y=150)

    # 提交按钮
    submit_button = tk.Button(window_edit, width=20, text="提交", fg='black', font=ft4, command=lambda: submit_info(wb,wb1,s1,s))
    submit_button.place(x=200, y=200)
def score_sort():
    print("test sort")
    f_score = open("score.in", "w")
    while 1:
        if q.empty():  #检验队列是否为空
            break
        #print(q.get())  #不为空输出队首且将队首删除
        f_score.write(str(q.get()))
        f_score.write('\n')
    f_score.close()

    f_score= open("score.in", "r", encoding='UTF-8')
    # code1=f1.read()
    scores =f_score.readlines()
    f_score.close()
    with open("name.in", "r", encoding='gbk') as f_name:  #open("score.in", "r", encoding='UTF-8') as f_score:
        names = f_name.readlines()
        # scores = f_score.readlines()

    if len(names) != len(scores):
        print("Error: Number of names and scores don't match!")
        return

    # Combine names, scores, and indices into a list of tuples
    data = [(names[i].strip(), int(float(scores[i].strip()))) for i in range(len(names))]

    # Implementing insertion sort and sort by scores
    for i in range(1, len(data)):
        key = data[i]
        j = i - 1
        while j >= 0 and data[j][1] < key[1]:
            data[j + 1] = data[j]
            j -= 1
        data[j + 1] = key

    # Assign ranks based on sorted order
    ranks = [1] * len(data)
    for i in range(1, len(data)):
        if data[i][1] == data[i - 1][1]:
            ranks[i] = ranks[i - 1]
        else:
            ranks[i] = ranks[i - 1] + 1

    # Combine sorted data with ranks
    # result = [f"姓名：{data[i][0]}，成绩：{data[i][1]}，排名：{ranks[i]}" for i in range(len(data))]
    # Combine sorted data with ranks, keeping only the first occurrence of each name
    result = []
    seen_names = set()
    for i in range(len(data)):
        if data[i][0] not in seen_names:
            result.append(f"姓名：{data[i][0]}，成绩：{data[i][1]}，排名：{ranks[i]}")
            seen_names.add(data[i][0])
    # Save sorted result to a file
    with open("sorted_scores.out", "w", encoding='UTF-8') as f_out:
        f_out.write('\n'.join(result))


import tkinter as tk


def menu(s, s1):
    # s为学生名单，s1为成绩单
    m = 0  # 行坐标
    y1 = 1  # 名字的起始列坐标
    y2 = 3  # 成绩的起始列坐标
    window_creat = tk.Toplevel(win)
    window_creat.configure(bg='LightBlue')
    window_creat.title("菜单界面")
    window_creat.geometry("600x500")

    # 新的样式
    background_color = 'LightYellow'
    text_color = 'Tomato'
    button_color = 'black'
    button_font = ('Arial', 12)

    Label(window_creat, text="请选择要执行的功能:", bg=background_color, fg=text_color, font=('Arial', 18)).pack(pady=10)
    Label(window_creat, text="提示：请依次执行1-4功能", bg='LightBlue', fg=text_color,font=('Arial',18)).pack(pady=10)
    Label(window_creat, text="进行功能5后，请重新依次执行各项功能", bg='LightBlue', fg=text_color,font=('Arial',18)).pack(pady=10)

    # 修改后的按钮样式
    button_1 = tk.Button(window_creat, text='1.读取学生名单', fg=button_color, font=button_font,
                         command=lambda: name(s))
    button_1.pack(pady=10)

    button_2 = tk.Button(window_creat, text='2.哈希冲突检测', fg=button_color, font=button_font,command=lambda: conflict(s, s1))
    button_2.pack(pady=10)

    button_3 = tk.Button(window_creat, text="3.成绩文件读取", fg=button_color, font=button_font,command=lambda: read(s, s1, m, y1, y2))
    button_3.pack(pady=10)

    button_4 = tk.Button(window_creat, text="4.预览学生名单输出", fg=button_color, font=button_font,
                         command=lambda: output(s, s1))
    button_4.pack(pady=10)

    button_5 = tk.Button(window_creat, text="5.添加/删除学生信息", fg=button_color, font=button_font,
                         command=lambda: edit(s, s1))
    button_5.pack(pady=10)

    button_6 = tk.Button(window_creat, text="6.退出程序", fg=button_color, font=button_font, command=win.quit)
    button_6.pack(pady=10)


if __name__ == "__main__":
    win = tk.Tk()
    win.geometry("500x400")  # 界面大小，中间是小写的x
    win.title("学生成绩匹配和管理系统")
    # win.configure(bg='LightBlue')
    # 加载图片
    image = Image.open("pic.gif")  # 替换为你的图片路径
    photo = ImageTk.PhotoImage(image)

    # 创建标签并设置图片
    label = tk.Label(win, image=photo)
    label.place(x=0, y=100, relwidth=1, relheight=1,anchor='nw')  # 使图片铺满整个窗口
    ft1 = tkFont.Font(family='Fixdsys', size=20, weight=tkFont.BOLD)
    ft2 = tkFont.Font(family='Fixdsys', size=10, weight=tkFont.BOLD)
    ft3 = tkFont.Font(family='Fixdsys', size=25, weight=tkFont.BOLD)
    ft4 = tkFont.Font(family='Fixdsys', size=15, weight=tkFont.BOLD)
    Label_first = tk.Label(win)
    Label_second = tk.Label(win)
    print("请输入要储存的学生名单文件名")
    print("请输入要打开的成绩文件名")
    s2="请依次输入学生名单的文件名和学生成绩单，如AA.xlsx和BB.xlsx"
    var1 = StringVar()
    var2 = StringVar()
    Label_first = tk.Label(win, text=s2, bg='LightGreen', fg='Tomato',font=ft2).place(x=55, y=0)
    Label_second = tk.Label(win, text="学生名单文件名", bg='LightGreen', fg='Maroon', justify='left').place(x=50, y=50)
    Label_third = tk.Label(win, text="学生成绩单文件名", bg='LightGreen', fg='Maroon', justify='left').place(x=50, y=100)
    text1 = tk.Entry(win, width=35, bg='White', textvariable=var1).place(x=170, y=50)
    text2 = tk.Entry(win, width=35, bg='White', textvariable=var2).place(x=170, y=100)

    button_First = tk.Button(win, width=5, text='Run!', fg='blue', command=lambda: menu(var1.get(),var2.get())).place(x=220, y=150)
    #button_Second = tk.Button(win, text='使用说明', activeforeground='red', width=10, height=2,command=create_window_uncover).place(x=230, y=140)
    win.mainloop()
    #menu()


